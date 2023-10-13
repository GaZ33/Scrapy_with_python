from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Abrindo o arquivo na variável
wb = load_workbook('pivot_table.xlsx')

# Selecionando a tabela que queremos usar
sheet = wb['Report']

# Atribuindo as colunas e linhas minimas e maximas
min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Criando a instânciar do barchart
barchar = BarChart()

# Selecionando os dados
data = Reference(sheet, min_col=min_col+1, max_col=max_col, min_row=min_row, max_row=max_row)
# Sekecionando as categorias dos dados
categories = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_row+1, max_row=max_row)

# Colocando os dados no objeto barchar que criamos
barchar.add_data(data, titles_from_data=True)
# Setando as categorias do barchar
barchar.set_categories(categories)

# Indicando onde começará o barchart
sheet.add_chart(barchar, "B12")

# Titúlo
barchar.title = "Sales"
# Estilo
barchar.style = 5

# Salvando ele
wb.save("barchart.xlsx")
