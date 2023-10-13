from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Abrindo o arquivo na variável
wb = load_workbook('barchart.xlsx')

# Selecionando a tabela que queremos usar
sheet = wb['Report']

# Atribuindo as colunas e linhas minimas e maximas
min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# # Precisamos selecionar a célula que desejamos trabalhar E colocando a formula na céulula
# sheet['E8'] = '=SUM(E6:E7)'
# # Colocando o style dessa célula
# sheet['E8'].style = 'Currency'

# Refazendo de outra maneira:
# Iterando toda as colunas das tabelas
for num in range(min_col+1, max_col+1):
    # Conseguindo a letra correspondente daquela coluna
    column = get_column_letter(num)
    # Formatando a função soma para aquela coluna
    linhas_soma = f'=SUM({column}{min_row+1}:{column}{max_row})'
    # Colocando a função soma na última linha+1 (para ficar de baixo da tabela)
    sheet[f'{column}{max_row+1}'] = f'{linhas_soma}'
    # Estilo
    sheet[f'{column}{max_row+1}'].style = 'Currency'

# Salvando o arquivo
wb.save("Report.xlsx")

